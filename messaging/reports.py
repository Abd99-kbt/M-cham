"""
نظام التقارير والإحصائيات
"""
from django.db.models import Count, Q
from django.utils import timezone
from datetime import timedelta, datetime
from .models import Message, MessageRecipient
from accounts.models import User, Department

class MessagingReports:
    """فئة إدارة تقارير الرسائل"""
    
    @staticmethod
    def get_message_statistics(start_date=None, end_date=None):
        """إحصائيات عامة للرسائل"""
        queryset = Message.objects.all()
        
        if start_date:
            queryset = queryset.filter(created_at__gte=start_date)
        if end_date:
            queryset = queryset.filter(created_at__lte=end_date)
        
        return {
            'total_messages': queryset.count(),
            'sent_messages': queryset.filter(status='SENT').count(),
            'draft_messages': queryset.filter(status='DRAFT').count(),
            'urgent_messages': queryset.filter(priority__in=['URGENT', 'CRITICAL']).count(),
            'confidential_messages': queryset.filter(confidentiality__in=['CONFIDENTIAL', 'TOP_SECRET']).count(),
            'messages_by_category': queryset.values('category__name').annotate(count=Count('id')),
            'messages_by_priority': queryset.values('priority').annotate(count=Count('id')),
            'messages_by_status': queryset.values('status').annotate(count=Count('id')),
        }
    
    @staticmethod
    def get_user_activity_report(user=None, days=30):
        """تقرير نشاط المستخدم"""
        end_date = timezone.now()
        start_date = end_date - timedelta(days=days)
        
        query_filter = {'created_at__gte': start_date, 'created_at__lte': end_date}
        if user:
            query_filter['sender'] = user
        
        sent_messages = Message.objects.filter(**query_filter)
        
        # الرسائل المستلمة
        received_filter = {'message__created_at__gte': start_date, 'message__created_at__lte': end_date}
        if user:
            received_filter['recipient'] = user
            
        received_messages = MessageRecipient.objects.filter(**received_filter)
        
        return {
            'user': user,
            'period_days': days,
            'sent_count': sent_messages.count(),
            'received_count': received_messages.count(),
            'sent_by_priority': sent_messages.values('priority').annotate(count=Count('id')),
            'sent_by_category': sent_messages.values('category__name').annotate(count=Count('id')),
            'read_rate': received_messages.filter(read_at__isnull=False).count() / max(received_messages.count(), 1) * 100,
            'avg_response_time': calculate_avg_response_time(received_messages),
        }
    
    @staticmethod
    def get_department_report(department=None, days=30):
        """تقرير نشاط القسم"""
        end_date = timezone.now()
        start_date = end_date - timedelta(days=days)
        
        query_filter = {'sender__department': department, 'created_at__gte': start_date}
        messages = Message.objects.filter(**query_filter)
        
        return {
            'department': department,
            'period_days': days,
            'total_messages': messages.count(),
            'messages_by_user': messages.values('sender__arabic_name').annotate(count=Count('id')),
            'messages_by_priority': messages.values('priority').annotate(count=Count('id')),
            'messages_by_category': messages.values('category__name').annotate(count=Count('id')),
            'urgent_messages': messages.filter(priority__in=['URGENT', 'CRITICAL']).count(),
        }
    
    @staticmethod
    def get_performance_metrics():
        """مؤشرات الأداء"""
        now = timezone.now()
        today = now.date()
        week_ago = now - timedelta(days=7)
        month_ago = now - timedelta(days=30)
        
        return {
            'today_messages': Message.objects.filter(created_at__date=today).count(),
            'week_messages': Message.objects.filter(created_at__gte=week_ago).count(),
            'month_messages': Message.objects.filter(created_at__gte=month_ago).count(),
            'unread_messages': MessageRecipient.objects.filter(read_at__isnull=True).count(),
            'pending_approvals': get_pending_approvals_count(),
            'system_health': get_system_health_status(),
        }
    
    @staticmethod
    def get_security_report(days=7):
        """تقرير الأمان"""
        from security.models import AuditLog, LoginAttempt
        
        end_date = timezone.now()
        start_date = end_date - timedelta(days=days)
        
        return {
            'period_days': days,
            'failed_login_attempts': LoginAttempt.objects.filter(
                is_successful=False,
                timestamp__gte=start_date
            ).count(),
            'successful_logins': LoginAttempt.objects.filter(
                is_successful=True,
                timestamp__gte=start_date
            ).count(),
            'security_violations': AuditLog.objects.filter(
                action_type='SECURITY_VIOLATION',
                timestamp__gte=start_date
            ).count(),
            'audit_logs_count': AuditLog.objects.filter(timestamp__gte=start_date).count(),
            'suspicious_activities': get_suspicious_activities(start_date),
        }

def calculate_avg_response_time(message_recipients):
    """حساب متوسط وقت الاستجابة"""
    total_time = timedelta()
    count = 0
    
    for recipient in message_recipients.filter(read_at__isnull=False):
        if recipient.message.sent_at and recipient.read_at:
            response_time = recipient.read_at - recipient.message.sent_at
            total_time += response_time
            count += 1
    
    if count > 0:
        avg_seconds = total_time.total_seconds() / count
        return round(avg_seconds / 3600, 2)  # بالساعات
    return 0

def get_pending_approvals_count():
    """عدد الموافقات المعلقة"""
    from workflows.models import ApprovalRequest
    return ApprovalRequest.objects.filter(status='PENDING').count()

def get_system_health_status():
    """حالة النظام"""
    try:
        # فحص قاعدة البيانات
        Message.objects.first()
        
        # فحص المساحة التخزينية
        import os
        import shutil
        total, used, free = shutil.disk_usage('.')
        free_percentage = (free / total) * 100
        
        if free_percentage < 10:
            return 'تحذير: مساحة التخزين منخفضة'
        elif free_percentage < 20:
            return 'تنبيه: مراقبة المساحة التخزينية'
        else:
            return 'جيد'
    except Exception as e:
        return f'خطأ: {str(e)}'

def get_suspicious_activities(start_date):
    """الأنشطة المشبوهة"""
    from security.models import AuditLog
    
    suspicious = []
    
    # محاولات دخول متعددة فاشلة
    from django.db.models import Count
    failed_attempts = AuditLog.objects.filter(
        action_type='FAILED_LOGIN',
        timestamp__gte=start_date
    ).values('user_ip').annotate(count=Count('id')).filter(count__gte=5)
    
    for attempt in failed_attempts:
        suspicious.append({
            'type': 'محاولات دخول متعددة فاشلة',
            'ip': attempt['user_ip'],
            'count': attempt['count']
        })
    
    return suspicious

def export_report_to_excel(report_data, filename):
    """تصدير التقرير إلى Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "تقرير الرسائل"
        
        # العنوان
        ws['A1'] = 'تقرير نظام الرسائل المصرفية'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # إضافة البيانات
        row = 3
        for key, value in report_data.items():
            ws[f'A{row}'] = str(key)
            ws[f'B{row}'] = str(value)
            row += 1
        
        wb.save(filename)
        return True
    except Exception as e:
        print(f"خطأ في تصدير Excel: {e}")
        return False

def export_report_to_pdf(report_data, filename):
    """تصدير التقرير إلى PDF"""
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        
        c = canvas.Canvas(filename, pagesize=A4)
        width, height = A4
        
        # العنوان
        c.setFont("Helvetica-Bold", 16)
        c.drawCentredText(width/2, height-50, "Banking Messaging System Report")
        
        # البيانات
        y = height - 100
        c.setFont("Helvetica", 12)
        
        for key, value in report_data.items():
            c.drawString(50, y, f"{key}: {value}")
            y -= 20
            
            if y < 50:  # صفحة جديدة
                c.showPage()
                y = height - 50
        
        c.save()
        return True
    except Exception as e:
        print(f"خطأ في تصدير PDF: {e}")
        return False